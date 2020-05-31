import openpyxl
import os
import ServerOperation
from ServerOperation import ServerData
import threading
from threading import Thread

AuthorityInfoPath = "Info\\Authority.xlsx"  # 存放管理员和经理密码的表格

class Administrator:  # 管理员

    def LogNameCheck(self):  # 检查用户登录信息是否正确
        password = input("请输入密码")
        return self.PasswordChange(password)

    def PasswordChange(self, password):  # 修改管理员账户密码
        print("密码正确",password)
        return True

    def GetTime(self):  # 获取当前系统时间
        pass

    def ShowRunInfo(self):  # 打印当前所有房间的信息
        print("我是管理员")

    def quit(self):  # 退出当前登录
        pass

    def run(self):  # 运行总函数
        print("管理员登录界面已启动")
        while True:
            if self.LogNameCheck() is True:
                while True:
                    self.ShowRunInfo()
                    self.quit()



class Manager(Administrator):  # 经理
    def PrintDetail(self):  # 打印当日指定时间的详单
        pass

    def run(self):
        print("经理登录界面已启动....")

class Waitor(Administrator):  # 前台
    def PrintBill(self):  # 打印账单
        pass

    def run(self):
        print("前台登录界面已启动,,,")


class Login:

    def ServerRun(self):
        run = ServerData()
        MonitorThread = Thread(target=self.FirstUIDeal)
        MonitorThread.setDaemon(True)
        MonitorThread.start()
        run.SocketStart()

    def LogNameCheck(self, id, password):  # 检查用户登录信息是否正确
        '''
        --------------------------------------------------------------------
        这里需要补充登录信息正确性检验
        -----------------------------------------------------------------------
        '''

        isExist = os.path.exists(AuthorityInfoPath)  # 检查登记表是否存在
        if isExist != 1:  # 还没有在主机端创建管理员信息表，就直接创建一个新的表
            os.makedirs("Info")
            table = openpyxl.Workbook()
            table.save(AuthorityInfoPath)
        excel = openpyxl.load_workbook(AuthorityInfoPath)
        data = excel.worksheets[0]
        #  passwordForCheck = data.cell(id - 1, 1).value  # 获取对应身份的密码
        passwordForCheck = "12345"  # 这里为了测试方便就直接用固定值
        print("###########", password)
        if id == 1 and password == passwordForCheck:
            print("Correct...")
            return True
        else:
            print("The password is WRONG...")
            return False

    def PasswordChange(self, id, password):  # 修改管理员账户密码
        '''
        --------------------------------------------------------------------
        这里需要补充密码修改功能
        -----------------------------------------------------------------------
        '''
        answer = self.LogNameCheck(id, password)
        if answer == False:
            return -1

        excel = openpyxl.load_workbook(AuthorityInfoPath)
        data = excel.worksheets[0]
        FirstPassword = input("Please input your new password :")
        SecondPassword = input("Please input your new password for second time:")
        if FirstPassword == SecondPassword:
            print("Change successful!...")
            # 修改管理员账户的密码

            data.cell(id - 1, 1, FirstPassword)
            excel.save(AuthorityInfoPath)
            return 0
        else:
            print("The new password is not same,Please Try again...")
            return -1

    def FirstUIDeal(self):
        waitor = Waitor()
        waitor_thread = threading.Thread(target=waitor.run)
        waitor_thread.setDaemon(True)
        waitor_thread.start()
        manager = Manager()
        manager_thread = threading.Thread(target=manager.run)
        manager_thread.setDaemon(True)
        manager_thread.start()
        administrator = Administrator()
        admin_thread = threading.Thread(target=administrator.run)
        admin_thread.setDaemon(True)
        admin_thread.start()


    '''
    def FirstUIDeal(self):  # 登录页，返回Flag表明输入的情况，1为成功，0为退出
        while True:  # 只要没有选择退出或者登陆成功，就一直在登录页面循环
            symbol = input("Choose your Identification: 1 for authority, 2 for header\n")
            #  print("symbol:", symbol)
            if symbol == "1":  # 管理员登录
                password = input("Please input the Password:")
                if self.LogNameCheck(1, password) == True:
                    print("登录成功，开始跳转...")
                    Flag = 1
                    self.user = Administrator()
                    break
            elif symbol == "2":  # 经理登录
                password = input("Please input the password:")
                if self.LogNameCheck(2, password) == True:
                    print("登录成功，开始跳转...")
                    Flag = 1
                    self.user = Manager()
                    break
            elif symbol == "3":  # 前台登录
                password = input("Please input the password:")
                if self.LogNameCheck(3, password) == True:
                    print("登录成功，开始跳转...")
                    Flag = 1
                    self.user = Waitor()
                    break
            else:  # 选择退出
                Flag = 0
                break
        return Flag
    '''
test = Login()
test.ServerRun()
